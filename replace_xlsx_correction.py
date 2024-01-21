import os.path

import openpyxl
from openpyxl.cell import MergedCell


def get_cell_value(sheet, cell):
    value = cell.value
    if value is None and isinstance(cell, MergedCell):
        value = get_merged_cell_value(sheet, cell)
    return value


def get_merged_cell_value(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value


def get_target_rows_from_sheet(sheet, target: str):
    is_target: bool = False
    rows = []
    for row in sheet.iter_rows(max_col=5):
        if row[0].value == target:
            is_target = True
        if row[0].value is None \
                and row[1].value is None \
                and row[2].value is None \
                and row[3].value is None \
                and row[4].value is None:
            is_target = False
        if is_target is True:

            cells = {
                'key': get_cell_value(sheet, row[1]),
                'value': get_cell_value(sheet, row[3]),
                'lan': get_cell_value(sheet, row[4])
            }
            # , get_cell_value(sheet, row[2]), get_cell_value(sheet, row[3]),
            # get_cell_value(sheet, row[4])
            rows.append(cells)
    return rows


def get_target_rows_from_workbook(workbook, target: str) -> dict:
    result = []
    for sheet in workbook.worksheets:
        rows = get_target_rows_from_sheet(sheet, target)
        result = result + rows

    # 创建空字典用于分类
    classified_dict = {}

    # 根据lan字段进行分类
    for obj in result:
        lan = obj['lan']
        if lan in classified_dict:
            classified_dict[lan].append(obj)
        else:
            classified_dict[lan] = [obj]
    return classified_dict


def update_localizable_string(localizable_string_file_path, string_list):
    if not os.path.exists(localizable_string_file_path):
        print('file not exist at path: ' + localizable_string_file_path)
        return

    content = ''
    with open(localizable_string_file_path, 'r') as file:
        lines = file.readlines()
        for line in lines:
            for i in string_list:
                tempString = '"'+i['key']+'"'
                if tempString in line:
                    line = f'"{i["key"]}" = "{i["value"]}";\n'
            content = content + line

    with open(localizable_string_file_path, 'w') as file:
        file.write(content)


def get_l10n_folder_name_with_key(key: str) -> str:

    if key.lower() == 'ko':
        return 'ko-KR.lproj'
    elif key.lower() == 'pt':
        return 'pt-PT.lproj'
    elif key.lower() == 'cn':
        return 'zh-Hans.lproj'
    elif key.lower() == 'tw':
        return 'zh-Hant.lproj'
    elif key.lower() == 'in':
        return 'id.lproj'
    else:
        return key.lower() + '.lproj'

#问 ： 有没有报错的情况  
# Press the green button in the gutter to run the script.
if __name__ == '__main__':

    xlsx_path = '/Users/haru.feng/Downloads/L10n_Bugs_for_5.15.10.xlsx'
    l10n_dir = '/Users/haru.feng/Documents/SourceCode/ep-new-release/mac-client/Zoom/Zoom/Zoom'
    workbook = openpyxl.load_workbook(xlsx_path)
    target: str = 'MAC'
    rows = get_target_rows_from_workbook(workbook, target)

    for lan, objs in rows.items():
        print(objs)
        folder_name = get_l10n_folder_name_with_key(lan)
        update_localizable_string(os.path.join(l10n_dir, folder_name, 'Localizable.strings'), objs)


        

